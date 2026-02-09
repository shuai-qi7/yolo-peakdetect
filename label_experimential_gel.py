import matplotlib.pyplot as plt
import os
import numpy as np
from openpyxl import Workbook, load_workbook
import time  # Allocate time for GUI rendering

# ========== Configure Matplotlib for Chinese Support + Force GUI Backend (Resolve Display Compatibility Issues) ==========
plt.rcParams['font.sans-serif'] = ['SimHei']  # SimHei font for Windows
plt.rcParams['axes.unicode_minus'] = False  # Fix negative sign display issue
# Force use of TkAgg backend (best compatibility, avoid display issues across different systems)
plt.switch_backend('TkAgg')
plt.ion()  # Enable interactive mode


# ========== Read 102 Columns of Data from XLSX ==========
def read_xlsx_columns(xlsx_path="sample_row_sums.xlsx"):
    """
    Read data from columns 1-102 of XLSX file, filter empty values, return list of valid data for each column
    :param xlsx_path: Path to XLSX file storing 102 columns of data
    :return: columns_data - List where each element is valid data (no empty values) for one column
    """
    columns_data = []
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        # Read columns 1-102 (strictly corresponding to sample IDs 1-102)
        for col_idx in range(1, 103):
            col_data = []
            # Start reading from row 2 (row 1 contains column headers)
            for row_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and cell_value != "":  # Filter empty values
                    col_data.append(float(cell_value))  # Convert to float to ensure proper calculations
            columns_data.append(col_data)
            print(f"Read column {col_idx} (sample {col_idx}) data, valid length: {len(col_data)}")

        wb.close()
        return columns_data
    except FileNotFoundError:
        print(f"Error: File {xlsx_path} not found, please check the path!")
        return []
    except Exception as e:
        print(f"Error reading XLSX: {e}")
        return []


'''
def find_nearest_point(x_arr, y_arr, click_x, click_y):
    """Find the point in original data closest to the clicked coordinates"""
    distances = np.sqrt((x_arr - click_x) ** 2 + (y_arr - click_y) ** 2)
    nearest_idx = np.argmin(distances)
    return x_arr[nearest_idx], y_arr[nearest_idx]
'''


def interactive_annotate_multi(sample_id, data_x, data_y):
    """
    Support continuous annotation: 1-Peak annotation 2-Overlapping peak annotation (direct annotation without confirmation),
    annotation numbering starts from 1 for each sample
    Compatible with older Matplotlib versions (remove plt.flush_events(), use fig.canvas.flush_events() instead)
    :param sample_id: Sample ID (directly use loop variable i, 1-102)
    :param data_x: Automatically generated X-axis data [1,2,...,len(data_y)]
    :param data_y: Column data read from XLSX
    """
    x_arr = np.array(data_x)
    y_arr = np.array(data_y)

    # Reset annotation sequence number to 0 for each sample (first peak/overlapping peak is numbered 1)
    peak_count = 0  # Peak annotation counter
    overlap_count = 0  # Overlapping peak annotation counter
    continue_annot = True

    # XLSX file path
    xlsx_path = 'multi_annot_records.xlsx'
    # Redefine header
    headers = [
        'Sample ID', 'Annotation Type', 'Annotation Number',
        'Peak Start X', 'Peak Start Y', 'Peak Max X', 'Peak Max Y', 'Peak End X', 'Peak End Y',
        'Overlap Peak Start X', 'Overlap Peak Start Y', 'Overlap Peak End X', 'Overlap Peak End Y',
        'Number of Overlapping Peaks', 'Total Area', 'Screenshot Filename', 'Annotation Date', 'Remarks'
    ]

    while continue_annot:
        # ========== Core Fix: Force cleanup of residual plots before each loop ==========
        plt.close('all')  # Close all residual plots to avoid resource occupation
        plt.ion()  # Reconfirm interactive mode (prevent accidental disable)

        # 1. Plot original curve (fully optimized display logic)
        fig, ax = plt.subplots(figsize=(16, 3))
        ax.plot(x_arr, y_arr, 'b-', label='Original Curve', linewidth=2)
        ax.set_xlabel('Swimlane Pixel Row Number')
        ax.set_ylabel('Pixel Sum')
        ax.set_ylim(0, 1.2 * max(y_arr))
        ax.set_title(
            f'Sample {sample_id} - Interactive Annotation (Current Peak Annotations: {peak_count} | Overlapping Peak Annotations: {overlap_count})')
        ax.grid(True)
        ax.legend()

        # ========== Rendering logic compatible with older versions (remove plt.flush_events()) ==========
        fig.canvas.draw()  # Draw plot content
        plt.tight_layout()
        plt.show(block=False)  # Non-blocking display
        plt.pause(0.1)  # Allocate 0.1 seconds for GUI rendering (critical!)

        # 2. Select annotation type
        print("\n===== Annotation Type Selection =====")
        print("1 - Peak Annotation (Need to click 3 points: Start → Peak Max → End)")
        print("2 - Overlapping Peak Annotation (Need to click 2 points: Start → End)")
        print("0 - Exit Annotation")
        annot_type = input("Enter number 0/1/2: ").strip()

        if annot_type == '0':
            print("Exit current sample annotation")
            plt.close(fig)
            break
        if annot_type not in ['1', '2']:
            print("Invalid input, please select 0/1/2!")
            plt.close(fig)
            continue

        # 3. Capture clicked points (update title and refresh)
        point_count = 3 if annot_type == '1' else 2
        prompt_text = f"Please click {point_count} points on the plot (Press Enter to confirm, timeout in 30 seconds):"
        print(prompt_text)
        ax.set_title(f'Sample {sample_id} - {prompt_text}')
        fig.canvas.draw()
        plt.pause(0.05)  # Allocate time for rendering after title refresh

        try:
            click_coords = plt.ginput(point_count, timeout=30)
            if len(click_coords) != point_count:
                print("Insufficient clicked points, skip current annotation")
                plt.close(fig)
                continue
        except:
            print("Click timeout/cancelled, skip current annotation")
            plt.close(fig)
            continue

        # 4. Match original data points
        click_points = []
        for i, (cx, cy) in enumerate(click_coords):
            label_x, label_y = int(cx), y_arr[int(cx)]
            click_points.append((label_x, label_y))

            # Set annotation point colors and labels
            if annot_type == '1':
                colors = ['green', 'red', 'blue']
                labels = ['Peak Start', 'Peak Max', 'Peak End']
            else:
                colors = ['orange', 'purple']
                labels = ['Overlap Peak Start', 'Overlap Peak End']
            ax.scatter(label_x, label_y, color=colors[i], s=150, zorder=5, label=labels[i])
            ax.annotate(labels[i], xy=(label_x, label_y),
                        xytext=(label_x + 0.2, label_y + 0.5),
                        arrowprops=dict(arrowstyle='->', color=colors[i]))

        # 5. Initialize annotation data
        annot_data = {
            'Sample ID': sample_id,
            'Annotation Type': '',
            'Annotation Number': 0,
            'Peak Start X': '', 'Peak Start Y': '', 'Peak Max X': '', 'Peak Max Y': '', 'Peak End X': '',
            'Peak End Y': '',
            'Overlap Peak Start X': '', 'Overlap Peak Start Y': '', 'Overlap Peak End X': '', 'Overlap Peak End Y': '',
            'Total Area': '', 'Screenshot Filename': '', 'Annotation Date': '20260130', 'Remarks': '',
            'Number of Overlapping Peaks': ''
        }

        if annot_type == '1':
            # Peak annotation logic
            peak_count += 1
            start_x, start_y = click_points[0]
            peak_x, peak_y = click_points[1]
            end_x, end_y = click_points[2]

            annot_data['Annotation Type'] = 'Peak Annotation'
            annot_data['Annotation Number'] = peak_count
            annot_data['Peak Start X'] = start_x
            annot_data['Peak Start Y'] = start_y
            annot_data['Peak Max X'] = peak_x
            annot_data['Peak Max Y'] = peak_y
            annot_data['Peak End X'] = end_x
            annot_data['Peak End Y'] = end_y

            annot_data[
                'Remarks'] = f'Peak {peak_count}: Start({start_x},{start_y}) Peak Max({peak_x},{peak_y}) End({end_x},{end_y})'

            # Highlight peak interval
            peak_mask = (x_arr >= min(start_x, end_x)) & (x_arr <= max(start_x, end_x))
            ax.plot(x_arr[peak_mask], y_arr[peak_mask], 'r--', linewidth=3, label=f'Peak {peak_count} Interval')
            screenshot_name = f'sample_{sample_id}_peak_{peak_count}_({start_x}-{peak_x}-{end_x}).png'

        else:
            # Overlapping peak annotation logic
            overlap_count += 1
            start_x, start_y = click_points[0]
            end_x, end_y = click_points[1]

            # Calculate total area
            sum_mask = (x_arr >= min(start_x, end_x)) & (x_arr <= max(start_x, end_x))
            sum_y = y_arr[sum_mask]
            total_area = np.sum(sum_y) if len(sum_y) > 0 else 0.0

            annot_data['Annotation Type'] = 'Overlapping Peak Annotation'
            annot_data['Annotation Number'] = overlap_count
            annot_data['Overlap Peak Start X'] = start_x
            annot_data['Overlap Peak Start Y'] = start_y
            annot_data['Overlap Peak End X'] = end_x
            annot_data['Overlap Peak End Y'] = end_y
            annot_data['Total Area'] = round(total_area, 4)

            # Input number of overlapping peaks
            print("\nPlease enter the number of overlapping peaks (integer):")
            overlap_num_str = input("Number of overlaps: ").strip()
            try:
                overlap_num = int(overlap_num_str)
                annot_data['Number of Overlapping Peaks'] = overlap_num
                overlap_note = f"Number of overlapping peaks: {overlap_num}"
            except:
                print("Invalid input for overlap count, recorded as empty")
                annot_data['Number of Overlapping Peaks'] = ''
                overlap_note = "Number of overlapping peaks: Invalid input"

            annot_data[
                'Remarks'] = f'Overlapping Peak {overlap_count}: Start({start_x},{start_y}) End({end_x},{end_y}) Total Area={total_area:.4f} {overlap_note}'

            # Highlight overlapping peak interval
            ax.plot(x_arr[sum_mask], y_arr[sum_mask], color='red', linestyle='--', linewidth=3,
                    label=f'Overlapping Peak {overlap_count} Interval')
            ax.text(0.7, 0.9, f'Overlapping Peak {overlap_count} Total Area={total_area:.4f}', transform=ax.transAxes,
                    bbox=dict(boxstyle='round', facecolor='yellow', alpha=0.5),
                    fontsize=12, fontweight='bold')
            screenshot_name = f'sample_{sample_id}_overlap_{overlap_count}_({start_x}-{end_x}).png'

        # 6. Save screenshot
        ax.legend(loc='best')
        ax.set_title(
            f'Sample {sample_id} - {annot_data["Annotation Type"]} {annot_data["Annotation Number"]} Completed')
        fig.canvas.draw()
        screenshot_path = os.path.join('annot_screenshots', screenshot_name)
        os.makedirs('annot_screenshots', exist_ok=True)
        plt.savefig(screenshot_path, dpi=300, bbox_inches='tight')
        annot_data['Screenshot Filename'] = screenshot_name

        # 7. Write to XLSX
        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Annotation Records'
            ws.append(headers)

        row_data = [annot_data[header] for header in headers]
        ws.append(row_data)
        wb.save(xlsx_path)
        wb.close()

        # 8. Close current plot and ask to continue
        plt.close(fig)  # Close current plot to avoid residue
        print(f"\n✅ Current annotation completed: {annot_data['Remarks']}")
        print(f"📸 Screenshot saved: {screenshot_name}")

        # Compatible with older versions: remove plt.flush_events(), direct input
        next_choice = input("\nContinue annotating current sample? (y/n): ").strip().lower()
        continue_annot = True if next_choice == 'y' else False

    # End of annotation: clean up all plots
    plt.close('all')
    total_annot = peak_count + overlap_count
    return f"\n===== Sample {sample_id} Annotation Completed =====\nTotal annotations: {peak_count} peaks, {overlap_count} overlapping peaks, {total_annot} total annotations\nAll data saved to multi_annot_records.xlsx"


# ========== Main Function: Process 102 Samples in Loop ==========
if __name__ == '__main__':
    # 1. Read data
    columns_data = read_xlsx_columns(xlsx_path="sample/sample_row_sums.xlsx")
    if len(columns_data) == 0:
        print("No valid column data read, program exited")
        exit()

    # 2. Process each sample in loop
    for sample_id in range(1, 103):
        col_idx = sample_id - 1  # List index starts from 0
        if col_idx >= len(columns_data):
            print(f"\n⚠️  No corresponding column data for sample {sample_id}, skip annotation")
            continue

        sample_y = columns_data[col_idx]
        if len(sample_y) == 0:
            print(f"\n⚠️  No valid data for sample {sample_id}, skip annotation")
            continue

        sample_x = list(range(1, len(sample_y) + 1))
        print(f"\n==================== Start Annotating Sample {sample_id} ====================")
        print(f"Sample {sample_id} data length: {len(sample_x)}")

        # Clean up all plots + refresh before switching samples
        plt.close('all')
        plt.ion()
        result = interactive_annotate_multi(
            sample_id=sample_id,
            data_x=sample_x,
            data_y=sample_y
        )
        print(result)

    # Program end: restore Matplotlib default state
    plt.ioff()
    plt.close('all')
    print("\n🎉 Annotation process for all 102 samples completed!")
